import asyncio
import logging
import os
import time
import tempfile
from datetime import datetime, timedelta
from threading import Thread
from typing import Dict, List

import pandas as pd
import schedule
from aiogram.types import BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from app.email_service.email_service import send_email_with_attachment
from app.inner_db.inner_db import applications_manager

logger = logging.getLogger(__name__)


# Цвета для отделов (обновленные)
DEPARTMENT_COLORS = {
    "ТЗ": PatternFill(
        start_color="A8E4A0", end_color="A8E4A0", fill_type="solid"
    ),  # Зеленый
    "ИМ": PatternFill(
        start_color="B0E0E6", end_color="B0E0E6", fill_type="solid"
    ),  # Голубой
    "🐕ТЗ": PatternFill(start_color="A8E4A0", end_color="A8E4A0", fill_type="solid"),
    "🐈ИМ": PatternFill(start_color="B0E0E6", end_color="B0E0E6", fill_type="solid"),
}


def create_excel_report(applications: List[Dict]) -> str:
    """Создает Excel-файл с отчетом по заявкам"""
    try:
        # Создаем папку для отчетов, если ее нет
        os.makedirs("/app/reports", exist_ok=True)

        # Подготавливаем данные для Excel
        data = []
        for app in applications:
            # Парсим текст заявки для извлечения данных
            app_text = app["application_text"]

            # Извлекаем данные из текста заявки
            date = extract_field(app_text, "Дата:")
            name = extract_field(app_text, "Имя:")
            department = extract_field(app_text, "Отдел:")
            contacts = extract_field(app_text, "Контакты:")
            comment = app.get("comment", "")
            status = app.get("status", "")

            # Получаем новые поля из базы данных
            db_contacts = app.get("contacts", "")
            db_name = app.get("name", "")
            db_purchaser_comment = app.get("purchaser_comment", "")

            # Извлекаем информацию о закупщике и сотруднике
            purchaser = extract_purchaser_info(app_text)  # Новая функция
            employee = extract_employee_info(app_text)  # Новая функция

            # Используем данные из базы, если они есть, иначе из текста
            final_contacts = db_contacts if db_contacts else contacts
            final_name = db_name if db_name else name

            # Если purchaser не найден в тексте, используем данные из базы
            if purchaser == "Не указано" and "purchaser" in app:
                purchaser = app.get("purchaser", "")

            # Если employee не найден в тексте, используем данные из базы
            if employee == "Не указано" and "employee" in app:
                employee = app.get("employee", "")

            # Извлекаем артикулы и названия книг
            articles, book_names = extract_books_info(app_text)

            data.append(
                {
                    "ID": app["id"],
                    "Дата в заявке": date,
                    "Имя": final_name,
                    "Отдел": department,
                    "Контакты": final_contacts,
                    "Закупщик": purchaser,
                    "Комментарий закупщика": db_purchaser_comment,
                    "Артикулы книг": articles,
                    "Названия книг": book_names,
                    "Сотрудник, внесший заявку": employee,
                    "Комментарий": comment,
                    "Статус": status,
                }
            )

        # Создаем DataFrame с правильным порядком колонок
        columns_order = [
            "ID",
            "Дата в заявке",
            "Имя",
            "Отдел",
            "Контакты",
            "Закупщик",
            "Комментарий закупщика",
            "Артикулы книг",
            "Названия книг",
            "Сотрудник, внесший заявку",
            "Комментарий",
            "Статус"
        ]

        df = pd.DataFrame(data)

        # Убедимся, что все колонки существуют (если каких-то нет, создадим пустые)
        for col in columns_order:
            if col not in df.columns:
                df[col] = ""

        # Переупорядочиваем колонки
        df = df[columns_order]

        # Генерируем имя файла с текущей датой
        current_date = datetime.now().strftime("%Y-%m-%d")
        temp_dir = tempfile.gettempdir()
        filename = os.path.join(temp_dir, f"Stol_zakazov_{current_date}.xlsx")

        # Создаем Excel файл с форматированием
        create_formatted_excel(df, filename)

        logger.info(f"✅ Excel-отчет создан: {filename}")
        return filename

    except Exception as e:
        logger.error(f"❌ Ошибка создания Excel-отчета: {e}")
        return None


def extract_purchaser_info(text: str) -> str:
    """Извлекает информацию о закупщике из текста заявки"""
    try:
        lines = text.split("\n")
        for line in lines:
            line_lower = line.lower()
            if "закупщик" in line_lower and ":" in line:
                return line.split(":", 1)[1].strip()
        return "Не указано"
    except:
        return "Ошибка извлечения"


def extract_employee_info(text: str) -> str:
    """Извлекает информацию о сотруднике, внесшем заявку"""
    try:
        lines = text.split("\n")
        for line in lines:
            line_lower = line.lower()
            if "сотрудник" in line_lower and ":" in line:
                return line.split(":", 1)[1].strip()
            elif "внесший" in line_lower and ":" in line:
                return line.split(":", 1)[1].strip()
        return "Не указано"
    except:
        return "Ошибка извлечения"


def create_formatted_excel(df: pd.DataFrame, filename: str):
    """Создает форматированный Excel файл с цветами для отделов"""
    # Создаем новую книгу
    wb = Workbook()

    # Удаляем дефолтный лист
    wb.remove(wb.active)

    # Создаем лист с заявками
    ws_applications = wb.create_sheet("Все заявки")

    # Добавляем данные из DataFrame
    for r_idx, row in enumerate(dataframe_to_rows(df,
                                                  index=False,
                                                  header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_applications.cell(row=r_idx, column=c_idx, value=value)

    # Применяем форматирование для отделов
    apply_department_formatting(ws_applications, df)

    # Автоматически подгоняем ширину столбцов
    auto_adjust_column_width(ws_applications)

    # Создаем лист со статистикой по статусам
    ws_status = wb.create_sheet("Статистика по статусам")
    status_stats = df["Статус"].value_counts().reset_index()
    status_stats.columns = ["Статус", "Количество"]

    for r_idx, row in enumerate(
        dataframe_to_rows(status_stats, index=False, header=True), 1
    ):
        for c_idx, value in enumerate(row, 1):
            ws_status.cell(row=r_idx, column=c_idx, value=value)

    # Создаем лист со статистикой по отделам
    ws_dept = wb.create_sheet("Статистика по отделам")
    dept_stats = df["Отдел"].value_counts().reset_index()
    dept_stats.columns = ["Отдел", "Количество"]

    for r_idx, row in enumerate(
        dataframe_to_rows(dept_stats, index=False, header=True), 1
    ):
        for c_idx, value in enumerate(row, 1):
            ws_dept.cell(row=r_idx, column=c_idx, value=value)

    # Сохраняем файл
    wb.save(filename)


def apply_department_formatting(worksheet, df: pd.DataFrame):
    """Применяет цветовое форматирование для отделов"""
    # Находим индекс столбца "Отдел"
    header_row = list(worksheet.iter_rows(
        min_row=1, max_row=1, values_only=True))[0]
    dept_col_idx = header_row.index("Отдел") + 1

    # Применяем цвета к строкам на основе отдела
    for row_idx in range(
        2, worksheet.max_row + 1
    ):  # Начинаем с 2 строки (после заголовка)
        department = worksheet.cell(row=row_idx, column=dept_col_idx).value

        if department in DEPARTMENT_COLORS:
            fill_color = DEPARTMENT_COLORS[department]
            # Закрашиваем всю строку
            for col_idx in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = fill_color


def auto_adjust_column_width(worksheet):
    """Автоматически подгоняет ширину столбцов"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width


def extract_field(text: str, field_name: str) -> str:
    """Извлекает значение поля из текста заявки"""
    try:
        lines = text.split("\n")
        for line in lines:
            if field_name in line:
                return line.split(field_name)[1].strip()
        return "Не указано"
    except:
        return "Ошибка извлечения"


def extract_books_info(text: str) -> tuple:
    """Извлекает артикулы и названия книг из текста заявки"""
    try:
        articles = []
        book_names = []

        lines = text.split("\n")
        for line in lines:
            line = line.strip()
            # Ищем строки с артикулами (предполагаем формат: "Артикул: значение")
            if line.startswith("Артикул:"):
                article = line.replace("Артикул:", "").strip()
                if article:
                    articles.append(article)

            # Ищем строки с названиями книг (предполагаем формат: "Название: значение")
            elif line.startswith("Название:"):
                book_name = line.replace("Название:", "").strip()
                if book_name:
                    book_names.append(book_name)

            # Альтернативный поиск по ключевым словам
            elif "артикул" in line.lower() and ":" in line:
                parts = line.split(":", 1)
                if len(parts) > 1:
                    article = parts[1].strip()
                    if article:
                        articles.append(article)

            elif "название" in line.lower() and ":" in line:
                parts = line.split(":", 1)
                if len(parts) > 1:
                    book_name = parts[1].strip()
                    if book_name:
                        book_names.append(book_name)

        # Если не нашли структурированных данных, пытаемся извлечь из общего текста
        if not articles and not book_names:
            articles, book_names = extract_from_unstructured_text(text)

        return ", ".join(articles) if articles else "Не указано", (
            ", ".join(book_names) if book_names else "Не указано"
        )

    except Exception as e:
        logger.error(f"Ошибка извлечения информации о книгах: {e}")
        return "Ошибка извлечения", "Ошибка извлечения"


def extract_from_unstructured_text(text: str) -> tuple:
    """Извлекает информацию о книгах из неструктурированного текста"""
    articles = []
    book_names = []

    # Простой парсинг для извлечения данных
    lines = text.split("\n")
    for line in lines:
        line = line.strip()

        # Пропускаем служебные строки
        if any(
            keyword in line
            for keyword in [
                "Заявка принята",
                "Дата:",
                "Контакты:",
                "Имя:",
                "Отдел:",
                "Комментарий:",
                "Статус заявки:",
                "Комментарий от закупщика:",
            ]
        ):
            continue

        # Если строка содержит цифры - возможно это артикул
        if any(char.isdigit() for char in line) and len(line) > 3:
            articles.append(line)
        # Иначе считаем это названием книги
        elif len(line) > 5:
            book_names.append(line)

    return articles, book_names


async def send_report_in_chat(message):
    """Отправляет отчет в чат пользователю"""
    try:
        applications = applications_manager.get_all_applications()

        if not applications:
            await message.answer("📝 Нет заявок для формирования отчета")
            return

        # Создаем Excel отчет
        excel_file = create_excel_report(applications)

        if excel_file and os.path.exists(excel_file):
            # Формируем статистику для сообщения
            status_counts = {}
            dept_counts = {}
            for app in applications:
                status = app["status"]
                status_counts[status] = status_counts.get(status, 0) + 1

                # Извлекаем отдел для статистики
                app_text = app["application_text"]
                department = extract_field(app_text, "Отдел:")
                dept_counts[department] = dept_counts.get(department, 0) + 1

            stats_message = "📊 Статистика отчета:\n"
            stats_message += f"📦 Всего заявок: {len(applications)}\n\n"
            stats_message += "📈 По статусам:\n"
            for status, count in status_counts.items():
                stats_message += f"• {status}: {count} заявок\n"

            stats_message += "\n🏢 По отделам:\n"
            for dept, count in dept_counts.items():
                stats_message += f"• {dept}: {count} заявок\n"

            await message.answer(stats_message)

            # Читаем файл в память и создаем BufferedInputFile
            with open(excel_file, "rb") as file:
                file_data = file.read()

            input_file = BufferedInputFile(
                file=file_data,
                filename=f"отчет_заявок_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            )

            # Отправляем файл
            await message.answer_document(
                document=input_file,
                caption=f"📎 Отчет по заявкам ({len(applications)} заявок)\n\n"
                f"🎨 Цветовая маркировка:\n"
                f"• 🟢 ТЗ - светло-зеленый (#A8E4A0)\n"
                f"• 🔵 ИМ - светло-голубой (#B0E0E6)",
            )

            # Удаляем временный файл
            os.remove(excel_file)

            logger.info(
                f"✅ Отчет отправлен в чат пользователю {message.from_user.id}")
        else:
            await message.answer("❌ Не удалось создать отчет")

    except Exception as e:
        await message.answer(f"❌ Ошибка при создании отчета: {e}")
        logger.error(f"Ошибка отправки отчета в чат: {e}")


async def send_report_to_email(email: str, user_id: int, user_name: str):
    """Отправляет отчет на указанный email"""
    try:
        applications = applications_manager.get_all_applications()

        if not applications:
            return False, "Нет заявок для формирования отчета"

        # Создаем Excel отчет
        excel_file = create_excel_report(applications)

        if excel_file and os.path.exists(excel_file):
            # Формируем текст email
            current_date = datetime.now().strftime("%d.%m.%Y %H:%M")
            subject = f"Отчет по заявкам от {current_date}"

            # Статистика для email
            status_counts = {}
            dept_counts = {}
            for app in applications:
                status = app["status"]
                status_counts[status] = status_counts.get(status, 0) + 1

                app_text = app["application_text"]
                department = extract_field(app_text, "Отдел:")
                dept_counts[department] = dept_counts.get(department, 0) + 1

            body = f"""
Отчет по заявкам

Дата формирования: {current_date}
Всего заявок: {len(applications)}
Запросил: {user_name} (ID: {user_id})

Статистика по статусам:
"""

            for status, count in status_counts.items():
                body += f"- {status}: {count} заявок\n"

            body += "\nСтатистика по отделам:\n"
            for dept, count in dept_counts.items():
                body += f"- {dept}: {count} заявок\n"

            body += f"\nЦветовая маркировка в файле:\n"
            body += f"- ТЗ: светло-зеленый (#A8E4A0)\n"
            body += f"- ИМ: светло-голубой (#B0E0E6)\n"
            body += f"\nЭтот отчет был запрошен через Telegram бота."

            # Отправляем email с отчетом
            email_sent = send_email_with_attachment(
                subject, body, excel_file, email)

            # Удаляем временный файл
            os.remove(excel_file)

            if email_sent:
                logger.info(
                    f"✅ Отчет отправлен на email {email} для пользователя {user_id}"
                )
                return True, f"Отчет отправлен на email: {email}"
            else:
                return False, "Не удалось отправить отчет на email"
        else:
            return False, "Не удалось создать отчет"

    except Exception as e:
        logger.error(f"Ошибка отправки отчета на email: {e}")
        return False, f"Ошибка: {e}"


def setup_weekly_report_scheduler():
    """Настраивает и запускает планировщик для еженедельных отчетов"""

    def schedule_thread():
        # Настраиваем расписание
        schedule.every().monday.at("09:00").do(send_weekly_report_job)

        logger.info(
            "📅 Планировщик еженедельных отчетов запущен (понедельник 9:00)")

        # Бесконечный цикл для выполнения запланированных задач
        while True:
            schedule.run_pending()
            time.sleep(60)  # Проверяем каждую минуту

    # Запускаем планировщик в отдельном потоке
    scheduler_thread = Thread(target=schedule_thread, daemon=True)
    scheduler_thread.start()


def send_weekly_report_job():
    """Задача для отправки еженедельного отчета"""
    try:
        logger.info("🔄 Запуск задачи отправки еженедельного отчета...")

        # Используем asyncio для запуска асинхронной функции
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

        # Получаем текущую дату для логирования
        current_date = datetime.now().strftime("%d.%m.%Y %H:%M")

        # Запускаем отправку отчета
        success, message = loop.run_until_complete(send_weekly_report())

        if success:
            logger.info(f"✅ Еженедельный отчет отправлен {current_date}")
        else:
            logger.error(f"❌ Ошибка отправки еженедельного отчета: {message}")

        loop.close()

    except Exception as e:
        logger.error(f"❌ Ошибка в задаче еженедельного отчета: {e}")


async def send_weekly_report():
    """Отправляет еженедельный отчет на email из конфигурации"""
    try:
        applications = applications_manager.get_all_applications()

        if not applications:
            return False, "Нет заявок для формирования отчета"

        # Создаем Excel отчет
        excel_file = create_excel_report(applications)

        if excel_file and os.path.exists(excel_file):
            # Получаем email из конфигурации
            from app.email_service.email_config import EMAIL_CONFIG

            # Основной email получателя
            main_email = EMAIL_CONFIG.get("email_to")

            # Дополнительные email из конфигурации
            additional_emails = EMAIL_CONFIG.get("weekly_report_emails", [])

            # Объединяем все email
            all_emails = (
                [main_email] + additional_emails if main_email else additional_emails
            )

            if not all_emails:
                return False, "Не указаны email для отправки отчета"

            # Формируем текст email
            current_date = datetime.now().strftime("%d.%m.%Y %H:%M")
            week_start = (datetime.now() - timedelta(days=7)
                          ).strftime("%d.%m.%Y")
            week_end = datetime.now().strftime("%d.%m.%Y")

            subject = (
                f"Еженедельный отчет по заявкам за период {week_start} - {week_end}"
            )

            # Статистика для email
            status_counts = {}
            dept_counts = {}
            new_this_week = 0

            for app in applications:
                status = app["status"]
                status_counts[status] = status_counts.get(status, 0) + 1

                app_text = app["application_text"]
                department = extract_field(app_text, "Отдел:")
                dept_counts[department] = dept_counts.get(department, 0) + 1

                # Считаем заявки за последнюю неделю
                app_date_str = extract_field(app_text, "Дата:")
                try:
                    app_date = datetime.strptime(
                        app_date_str, "%d.%m.%Y %H:%M")
                    if datetime.now() - app_date <= timedelta(days=7):
                        new_this_week += 1
                except:
                    pass

            body = f"""
ЕЖЕНЕДЕЛЬНЫЙ ОТЧЕТ ПО ЗАЯВКАМ

Период: {week_start} - {week_end}
Дата формирования: {current_date}

ОБЩАЯ СТАТИСТИКА:
Всего заявок в системе: {len(applications)}
Новых заявок за неделю: {new_this_week}

СТАТИСТИКА ПО СТАТУСАМ:
"""

            for status, count in status_counts.items():
                body += f"- {status}: {count} заявок\n"

            body += "\nСТАТИСТИКА ПО ОТДЕЛАМ:\n"
            for dept, count in dept_counts.items():
                body += f"- {dept}: {count} заявок\n"

            body += f"\nЦВЕТОВАЯ МАРКИРОВКА В ФАЙЛЕ:\n"
            body += f"- ТЗ: светло-зеленый (#A8E4A0)\n"
            body += f"- ИМ: светло-голубой (#B0E0E6)\n"
            body += f"\nЭто автоматический еженедельный отчет из Telegram бота."

            # Отправляем email всем получателям
            email_results = []
            for email in all_emails:
                if email:  # Проверяем, что email не пустой
                    email_sent = send_email_with_attachment(
                        subject, body, excel_file, email
                    )
                    email_results.append((email, email_sent))

            # Удаляем временный файл
            os.remove(excel_file)

            # Анализируем результаты отправки
            successful_emails = [email for email,
                                 sent in email_results if sent]
            failed_emails = [email for email,
                             sent in email_results if not sent]

            if successful_emails:
                logger.info(
                    f"✅ Отчет отправлен на {len(successful_emails)} email: {', '.join(successful_emails)}"
                )

            if failed_emails:
                logger.error(
                    f"❌ Не удалось отправить отчет на {len(failed_emails)} email: {', '.join(failed_emails)}"
                )

            if successful_emails:
                return True, f"Отчет отправлен на {len(successful_emails)} email"
            else:
                return False, "Не удалось отправить отчет ни на один email"

        else:
            return False, "Не удалось создать отчет"

    except Exception as e:
        logger.error(f"Ошибка отправки еженедельного отчета: {e}")
        return False, f"Ошибка: {e}"

